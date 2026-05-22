import json
import os
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Rutas ──────────────────────────────────────────────────────────────────────
# El JSON lo guarda Godot en su carpeta de usuario:
# Windows: C:\Users\<usuario>\AppData\Roaming\Godot\app_userdata\Laberinto\
GODOT_USER = Path(r"C:/Users/maure/Desktop/finalproyecto/Laberinto-Perdido-main/Laberinto-Perdido-main")
JSON_PATH  = GODOT_USER / "reporte_partida.json"
EXCEL_PATH = Path(r"C:/Users/maure/Desktop/finalproyecto/Laberinto-Perdido-main/Laberinto-Perdido-main/laberinto_scores.xlsx")

# Permitir override por argumento: python registrar_partida.py "C:/ruta/reporte.json"
if len(sys.argv) > 1:
    JSON_PATH = Path(sys.argv[1])

# ── Colores ────────────────────────────────────────────────────────────────────
COLOR_HEADER   = "1F3864"   # azul oscuro
COLOR_WIN      = "C6EFCE"   # verde claro
COLOR_LOSE     = "FFCCCC"   # rojo claro
COLOR_TOP      = "FFD700"   # dorado para top 3
FONT_WHITE     = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_NORMAL    = Font(name="Arial", size=10)
FONT_BOLD      = Font(name="Arial", bold=True, size=10)
ALIGN_CENTER   = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT     = Alignment(horizontal="left",   vertical="center")

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = ["#", "Jugador", "Estado", "Tiempo", "Segundos", "Monedas", "Fecha"]
WIDTHS   = [5,   20,        18,       12,        12,          10,        20   ]

def leer_json() -> dict:
    if not JSON_PATH.exists():
        sys.exit(f"❌  No se encontró el archivo JSON en:\n    {JSON_PATH}\n"
                 f"    Asegúrate de haber terminado una partida.")
    with open(JSON_PATH, encoding="utf-8") as f:
        return json.load(f)

def crear_excel_nuevo(wb: openpyxl.Workbook):
    """Crea las dos hojas con sus encabezados."""
    # Hoja 1: Registro completo
    ws = wb.active
    ws.title = "Partidas"
    ws.row_dimensions[1].height = 28

    for col, (titulo, ancho) in enumerate(zip(COLUMNS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font    = FONT_WHITE
        cell.fill    = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = ALIGN_CENTER
        cell.border  = BORDER
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws.freeze_panes = "A2"

    # Hoja 2: Top Tiempos
    ws2 = wb.create_sheet("🏆 Mejores Tiempos")
    ws2.row_dimensions[1].height = 28
    top_cols   = ["#", "Jugador", "Tiempo", "Monedas", "Fecha"]
    top_widths = [5,   20,        12,        10,         20    ]
    for col, (titulo, ancho) in enumerate(zip(top_cols, top_widths), start=1):
        cell = ws2.cell(row=1, column=col, value=titulo)
        cell.font      = FONT_WHITE
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER
        ws2.column_dimensions[get_column_letter(col)].width = ancho
    ws2.freeze_panes = "A2"

def agregar_partida(ws, datos: dict) -> int:
    """Agrega una fila al registro y devuelve el número de fila."""
    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(2, 1).value is None:
        next_row = 2  # Hoja vacía

    num       = next_row - 1
    completado = datos.get("completado", False)
    estado    = "✅ Completó" if completado else "❌ No completó"
    color_fill = PatternFill("solid", fgColor=COLOR_WIN if completado else COLOR_LOSE)

    valores = [
        num,
        datos.get("jugador", "?"),
        estado,
        datos.get("tiempo_formato", "00:00.00"),
        datos.get("tiempo_segundos", 0),
        datos.get("monedas", 0),
        datos.get("fecha", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    aligns  = [ALIGN_CENTER, ALIGN_LEFT, ALIGN_CENTER,
               ALIGN_CENTER, ALIGN_CENTER, ALIGN_CENTER, ALIGN_CENTER]

    for col, (val, aln) in enumerate(zip(valores, aligns), start=1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font      = FONT_NORMAL
        cell.fill      = color_fill
        cell.alignment = aln
        cell.border    = BORDER

    ws.row_dimensions[next_row].height = 20
    return next_row

def actualizar_top(ws_top, ws_partidas):
    """Reconstruye la hoja de mejores tiempos leyendo las partidas completadas."""
    # Borrar filas anteriores (dejar encabezado)
    for row in ws_top.iter_rows(min_row=2, max_row=ws_top.max_row):
        for cell in row:
            cell.value = None
            cell.fill  = PatternFill(fill_type=None)
            cell.font  = FONT_NORMAL

    # Recolectar partidas completadas
    completadas = []
    for row in ws_partidas.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        estado = str(row[2])
        if "Completó" in estado and "No" not in estado:
            completadas.append({
                "jugador":  row[1],
                "tiempo":   row[3],
                "segundos": float(row[4]) if row[4] else 99999,
                "monedas":  row[5],
                "fecha":    row[6],
            })

    # Ordenar por tiempo (menor = mejor)
    completadas.sort(key=lambda x: x["segundos"])

    medallas = ["🥇", "🥈", "🥉"]
    for i, p in enumerate(completadas):
        fila = i + 2
        es_top3 = i < 3
        color = PatternFill("solid", fgColor=COLOR_TOP) if es_top3 else PatternFill(fill_type=None)
        fuente = FONT_BOLD if es_top3 else FONT_NORMAL

        puesto = medallas[i] if i < 3 else str(i + 1)
        valores = [puesto, p["jugador"], p["tiempo"], p["monedas"], p["fecha"]]
        aligns  = [ALIGN_CENTER, ALIGN_LEFT, ALIGN_CENTER, ALIGN_CENTER, ALIGN_CENTER]

        for col, (val, aln) in enumerate(zip(valores, aligns), start=1):
            cell = ws_top.cell(row=fila, column=col, value=val)
            cell.font      = fuente
            cell.fill      = color
            cell.alignment = aln
            cell.border    = BORDER
        ws_top.row_dimensions[fila].height = 20

def main():
    datos = leer_json()

    # Cargar o crear Excel
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws_partidas = wb["Partidas"]
        ws_top      = wb["🏆 Mejores Tiempos"]
    else:
        wb = openpyxl.Workbook()
        crear_excel_nuevo(wb)
        ws_partidas = wb["Partidas"]
        ws_top      = wb["🏆 Mejores Tiempos"]
        print(f"📊 Excel nuevo creado en: {EXCEL_PATH}")

    agregar_partida(ws_partidas, datos)
    actualizar_top(ws_top, ws_partidas)
    wb.save(EXCEL_PATH)

    print("╔══════════════════════════════════════╗")
    print("║       REPORTE FINAL DE PARTIDA        ║")
    print("╠══════════════════════════════════════╣")
    print("║  Jugador : %s" % datos.get("jugador", "?"))
    icono = "✅" if datos.get("completado") else "❌"
    estado = "COMPLETÓ EL LABERINTO" if datos.get("completado") else "NO COMPLETÓ"
    print("║  Estado  : %s %s" % (icono, estado))
    print("║  Tiempo  : %s" % datos.get("tiempo_formato", "?"))
    print("║  Monedas : %d" % datos.get("monedas", 0))
    print("║  Fecha   : %s" % datos.get("fecha", "?"))
    print("╚══════════════════════════════════════╝")
    print(f"\n✅ Registrado en Excel → {EXCEL_PATH}")

if __name__ == "__main__":
    main()
